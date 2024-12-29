import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime
from .models import Capacity,Free,CapacitySection,TypePoints,CourseYear
from AcademicInfoApp.models import Year,Section,Course,Language
from UsersApp.models import Professor
from .forms import UploadForm
from django.shortcuts import render, redirect,get_object_or_404
from django.contrib import messages
from django.db.models import Sum
import re

def sanitize_sheet_title(title):
    # Invalid characters for Excel sheet names
    invalid_chars = r'[\\/:*?\"<>|]'
    return re.sub(invalid_chars, '_', title)

## CAPACITYPROFESSORS EXCEL

def generate_capacityprofessor_excel(request,year_id):
    try:
        # Validate and fetch the selected year
        try:
            selected_year = Year.objects.get(pk=int(year_id))
        except (Year.DoesNotExist, ValueError, TypeError):
            messages.error(request, f"Selecciona un curs acadèmic que existeixi: {str(e)}")
            return redirect("capacityprofessor_list")


        # Fetch all sections and professors for the selected year
        all_sections = Section.objects.all()
        professor_data = []

        capacities = Capacity.objects.filter(Year_id=selected_year.idYear).order_by('Professor__family_name')

        for capacity in capacities:
            professor = capacity.Professor
            free_points = Free.objects.filter(Professor=professor, Year=selected_year).aggregate(free_points=Sum('PointsFree'))['free_points'] or 0
            section_points = [(section.LetterSection, 0) for section in all_sections]

            for section_entry in CapacitySection.objects.filter(Professor=professor, Year=selected_year):
                section_letter = section_entry.Section.LetterSection
                section_points = [(letter, section_entry.Points if letter == section_letter else points) for letter, points in section_points]

            section_points_sum = sum(points for _, points in section_points)
            balance = capacity.Points - free_points - section_points_sum

            professor_data.append({
                'id': professor.idProfessor,
                'name': professor.name,
                'family_name': professor.family_name,
                'total_points': capacity.Points,
                'free_points': free_points,
                'sections': dict(section_points),
                'balance': balance,
            })

        # Generate Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sanitize_sheet_title(f"Capacitat_docent_{selected_year.Year}")

        # Header row
        headers = ['ID del Professor', 'Nom', 'Cognom', 'Punts totals', "Punts d'alliberació"] + [section.LetterSection for section in all_sections] + ['Balanç']
        sheet.append(headers)

        # Data rows
        for data in professor_data:
            row = [
                data['id'],
                data['name'],
                data['family_name'],
                data['total_points'],
                data['free_points'],
            ]
            row.extend([data['sections'].get(section.LetterSection, 0) for section in all_sections])
            row.append(data['balance'])
            sheet.append(row)

        # Adjust column widths
        for col in sheet.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value) + 2
            sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length

        # Create response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="capacitat_docent_{selected_year.Year}.xlsx"'
        workbook.save(response)

        return response
    except Exception as e:
        messages.error(request, f"Error al generar el fitxer Excel: {str(e)}")
        return redirect("capacityprofessor_list")

def upload_capacityprofessor_excel(request):
    return 
#     if request.method == "POST":
#         form = UploadForm(request.POST, request.FILES)
#         if form.is_valid():
#             excel_file = request.FILES["file"]
#             error_occurred = False
#             try:
#                 workbook = openpyxl.load_workbook(excel_file)
#                 sheet = workbook.active

#                 expected_columns =["ID del Grau","Nom del Grau","Codi del Grau","Escola","Està Actiu"]
#                 header_row = [cell.value for cell in sheet[1]]

#                 # Check columns
#                 if not all(col in header_row for col in expected_columns):
#                     messages.error(request, "El fitxer no té les columnes esperades. Comprova que el fitxer segueixi les dades com el excel a descarregar.")
#                     return render(request, "degree_upload_form.html", {"form": form})


#                 #skip header
#                 for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
#                     id_degree = row[0]
#                     name_degree = row[1]
#                     code_degree = row[2]
#                     school_name = row[3]
#                     is_active = row[4]

#                     if not all([id_degree, name_degree, code_degree, school_name]):
#                         messages.warning(
#                             request, f"Fila {row_num} no s'ha processat correctament: informació incompleta."
#                         )
#                         error_occurred=True
#                         continue

#                     if not isinstance(id_degree, (int, float)) or not str(id_degree).isdigit():
#                         messages.warning(
#                             request, f"Fila {row_num} no s'ha processat: l'ID de titulació ha de ser un número."
#                         )
#                         error_occurred = True
#                         continue

#                     if not isinstance(code_degree, (int, float)) or not str(code_degree).isdigit():
#                         messages.warning(
#                             request, f"Fila {row_num} no s'ha processat: el codi de titulació ha de ser un número."
#                         )
#                         error_occurred = True
#                         continue

#                     if is_active not in ['Si', 'No']:
#                         messages.warning(
#                             request, f"Fila {row_num} no s'ha processat: l'estat d'activitat ha de ser 'Si' o 'No'."
#                         )
#                         error_occurred = True
#                         continue

#                     try:
#                         school = Free.objects.get(NameSchool=school_name)
#                     except School.DoesNotExist:
#                         messages.warning(
#                             request, f"Fila {row_num} no s'ha processat: l'escola '{school_name}' no existeix."
#                         )
#                         error_occurred=True
#                         continue

#                     #If exists
#                     existing_degree = Degree.objects.filter(idDegree=id_degree).first()

#                     if existing_degree:
#                         if Degree.objects.filter(NameDegree=name_degree, School=school).exclude(idDegree=existing_degree.idDegree).exists():
#                             messages.warning(
#                                 request, f"Fila {row_num} no s'ha processat: la combinació de titulació i escola ja existeix."
#                             )
#                             error_occurred = True
#                             continue
#                         if Degree.objects.filter(CodeDegree=code_degree).exclude(CodeDegree=existing_degree.CodeDegree).exists():
#                             messages.warning(
#                                 request, f"Fila {row_num} no s'ha processat: el codi de la titulació ja existeix."
#                             )
#                             error_occurred = True
#                             continue
#                         # Update the existing degree
#                         existing_degree.NameDegree = name_degree
#                         existing_degree.CodeDegree = code_degree
#                         existing_degree.School = school
#                         existing_degree.isActive = is_active.lower() == "Si" if isinstance(is_active, str) else bool(is_active)
#                         existing_degree.save()

#                     #Does not exist - create
#                     else:
#                         if Degree.objects.filter(CodeDegree=code_degree).exists():
#                             messages.warning(
#                                 request, f"Fila {row_num} no s'ha processat: el codi de la titulació ja existeix."
#                             )
#                             error_occurred = True
#                             continue

#                         if Degree.objects.filter(NameDegree=name_degree, School=school).exists():
#                             messages.warning(
#                                 request, f"Fila {row_num} no s'ha processat: la combinació de titulació i escola ja existeix."
#                             )
#                             error_occurred = True
#                             continue

#                         # Create a new degree
#                         degree = Degree.objects.create(
#                             idDegree=id_degree,
#                             NameDegree=name_degree,
#                             CodeDegree=code_degree,
#                             School=school,
#                             isActive=is_active.lower() == "Si" if isinstance(is_active, str) else bool(is_active),
#                         )
                
#                 if not error_occurred:
#                     messages.success(request, "Els graus s'han actualitzat correctament.")
#                     return redirect('degree_list')

#             except Exception as e:
#                 messages.error(request, f"Error en processar el fitxer: {e}")
#                 error_occurred=True

#             return render(request, "degree_upload_form.html", {"form": form})

#     else:
#         form = UploadForm()

#     return render(request, "degree_upload_form.html", {"form": form})
