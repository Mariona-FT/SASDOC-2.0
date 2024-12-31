# services.py
from django.contrib.auth.hashers import make_password
from django.contrib import messages
import pandas as pd
from UsersApp.models import CustomUser, Professor,ProfessorField,ProfessorLanguage,TypeProfessor
import openpyxl
from openpyxl.styles import Alignment
from django.http import HttpResponse
from datetime import datetime
from .forms import UploadForm
from django.shortcuts import render, redirect,get_object_or_404
from django.contrib import messages

def generate_professor_excel(request):
    try:
        # Create a new Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Professorat"

        headers = [
            ("ID del Professor", 15),
            ("Username",20),
            ("Nom", 20),
            ("Cognoms", 30),
            ("Correu",30),
            ("Contracte actual",25),
            ("Camps de coneixament",30),
            ("Idiomes",30),
            ("Està Actiu", 15),
            ("Comentari cap de secció", 50),
            ("Descripció direcció", 50),
        ]

        # Add headers to the first row
        for col_num, (header, width) in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

        # Write the data for each professor
        professors = Professor.objects.all()

        for row_num, professor in enumerate(professors, start=2):
            # Get related fields, languages, and contract
            fields = ProfessorField.objects.filter(Professor=professor)
            languages = ProfessorLanguage.objects.filter(Professor=professor)
            contract = professor.current_contract

            # Join multiple fields/languages with commas, handle nulls
            fields_str = ", ".join(field.Field.NameField for field in fields) if fields.exists() else "Sense Camps"
            languages_str = ", ".join(lang.Language.Language for lang in languages) if languages.exists() else "Sense Idiomes"
            contract_str = contract.NameContract if contract else "Sense Contracte"

            # Populate the row
            sheet.cell(row=row_num, column=1, value=professor.idProfessor).alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=2, value=professor.user.username if professor.user else "Sense Usuari").alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=3, value=professor.name).alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=4, value=professor.family_name).alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=5, value=professor.email).alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=6, value=contract_str).alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=7, value=fields_str).alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=8, value=languages_str).alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=9, value="Sí" if professor.isActive == "yes" else "No").alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=10, value=professor.comment).alignment = Alignment(horizontal="left")
            sheet.cell(row=row_num, column=11, value=professor.description).alignment = Alignment(horizontal="left")

        # Generate a timestamped file name
        current_datetime = datetime.now().strftime("%d%m%Y_%H%M%S")
        filename = f"professorat_{current_datetime}.xlsx"

        # Set up the HTTP response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        workbook.save(response)

        return response
    except Exception as e:
        messages.error(request, f"Error al generar el fitxer Excel: {str(e)}")
        return redirect("usersapp:professor_list")

def upload_professor_excel(request):
    pass
    # if request.method == "POST":
    #     form = UploadForm(request.POST, request.FILES)
    #     if form.is_valid():
    #         excel_file = request.FILES["file"]
    #         error_occurred = False
    #         try:
    #             workbook = openpyxl.load_workbook(excel_file)
    #             sheet = workbook.active

    #             expected_columns =["ID del Grau","Nom del Grau","Codi del Grau","Escola","Està Actiu"]
    #             header_row = [cell.value for cell in sheet[1]]

    #             # Check columns
    #             if not all(col in header_row for col in expected_columns):
    #                 messages.error(request, "El fitxer no té les columnes esperades. Comprova que el fitxer segueixi les dades com el excel a descarregar.")
    #                 return render(request, "degree_upload_form.html", {"form": form})


    #             #skip header
    #             for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    #                 id_degree = row[0]
    #                 name_degree = row[1]
    #                 code_degree = row[2]
    #                 school_name = row[3]
    #                 is_active = row[4]

    #                 if not all([id_degree, name_degree, code_degree, school_name]):
    #                     messages.warning(
    #                         request, f"Fila {row_num} no s'ha processat correctament: informació incompleta."
    #                     )
    #                     error_occurred=True
    #                     continue

    #                 if not isinstance(id_degree, (int, float)) or not str(id_degree).isdigit():
    #                     messages.warning(
    #                         request, f"Fila {row_num} no s'ha processat: l'ID de titulació ha de ser un número."
    #                     )
    #                     error_occurred = True
    #                     continue

    #                 if not isinstance(code_degree, (int, float)) or not str(code_degree).isdigit():
    #                     messages.warning(
    #                         request, f"Fila {row_num} no s'ha processat: el codi de titulació ha de ser un número."
    #                     )
    #                     error_occurred = True
    #                     continue

    #                 if is_active not in ['Si', 'No']:
    #                     messages.warning(
    #                         request, f"Fila {row_num} no s'ha processat: l'estat d'activitat ha de ser 'Si' o 'No'."
    #                     )
    #                     error_occurred = True
    #                     continue

    #                 try:
    #                     school = School.objects.get(NameSchool=school_name)
    #                 except School.DoesNotExist:
    #                     messages.warning(
    #                         request, f"Fila {row_num} no s'ha processat: l'escola '{school_name}' no existeix."
    #                     )
    #                     error_occurred=True
    #                     continue

    #                 #If exists
    #                 existing_degree = Degree.objects.filter(idDegree=id_degree).first()

    #                 if existing_degree:
    #                     if Degree.objects.filter(NameDegree=name_degree, School=school).exclude(idDegree=existing_degree.idDegree).exists():
    #                         messages.warning(
    #                             request, f"Fila {row_num} no s'ha processat: la combinació de titulació i escola ja existeix."
    #                         )
    #                         error_occurred = True
    #                         continue
    #                     if Degree.objects.filter(CodeDegree=code_degree).exclude(CodeDegree=existing_degree.CodeDegree).exists():
    #                         messages.warning(
    #                             request, f"Fila {row_num} no s'ha processat: el codi de la titulació ja existeix."
    #                         )
    #                         error_occurred = True
    #                         continue
    #                     # Update the existing degree
    #                     existing_degree.NameDegree = name_degree
    #                     existing_degree.CodeDegree = code_degree
    #                     existing_degree.School = school
    #                     existing_degree.isActive = is_active.lower() == "Si" if isinstance(is_active, str) else bool(is_active)
    #                     existing_degree.save()

    #                 #Does not exist - create
    #                 else:
    #                     if Degree.objects.filter(CodeDegree=code_degree).exists():
    #                         messages.warning(
    #                             request, f"Fila {row_num} no s'ha processat: el codi de la titulació ja existeix."
    #                         )
    #                         error_occurred = True
    #                         continue

    #                     if Degree.objects.filter(NameDegree=name_degree, School=school).exists():
    #                         messages.warning(
    #                             request, f"Fila {row_num} no s'ha processat: la combinació de titulació i escola ja existeix."
    #                         )
    #                         error_occurred = True
    #                         continue

    #                     # Create a new degree
    #                     degree = Degree.objects.create(
    #                         idDegree=id_degree,
    #                         NameDegree=name_degree,
    #                         CodeDegree=code_degree,
    #                         School=school,
    #                         isActive=is_active.lower() == "Si" if isinstance(is_active, str) else bool(is_active),
    #                     )
                
    #             if not error_occurred:
    #                 messages.success(request, "Els graus s'han actualitzat correctament.")
    #                 return redirect('degree_list')

    #         except Exception as e:
    #             messages.error(request, f"Error en processar el fitxer: {e}")
    #             error_occurred=True

    #         return render(request, "degree_upload_form.html", {"form": form})

    # else:
    #     form = UploadForm()

    # return render(request, "degree_upload_form.html", {"form": form})