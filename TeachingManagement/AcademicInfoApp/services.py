import openpyxl
from openpyxl.styles import Alignment
from django.http import HttpResponse
from datetime import datetime
from .models import Degree,School
from .forms import DegreeUploadForm
from django.shortcuts import render, redirect,get_object_or_404
from django.contrib import messages

## DEGREE EXCEL

def generate_degree_excel(request):
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Graus"

    headers = [
        ("ID del Grau", 15),
        ("Nom del Grau", 30),
        ("Codi del Grau", 15),
        ("Escola", 30),
        ("Està Actiu", 15),
    ]

    # headers to the first row
    for col_num, (header, width) in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

    # Write the data for each Degree
    degrees = Degree.objects.all()
    for row_num, degree in enumerate(degrees, start=2):
        sheet.cell(row=row_num, column=1, value=degree.idDegree)
        sheet.cell(row=row_num, column=2, value=degree.NameDegree)
        sheet.cell(row=row_num, column=3, value=degree.CodeDegree)
        sheet.cell(row=row_num, column=4, value=degree.School.NameSchool if degree.School else "Sense Escola")
        sheet.cell(row=row_num, column=5, value="Actiu" if degree.isActive else "Inactiu")

    # Generate a timestamped file name
    current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"graus_{current_datetime}.xlsx"

    # Set up the HTTP response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    workbook.save(response)

    return response

def upload_degree_excel(request):
    if request.method == "POST":
        form = DegreeUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES["file"]
            error_occurred = False
            try:
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active

                expected_columns =["ID del Grau","Nom del Grau","Codi del Grau","Escola","Està Actiu"]
                header_row = [cell.value for cell in sheet[1]]

                # Check columns
                if not all(col in header_row for col in expected_columns):
                    messages.error(request, "El fitxer no té les columnes esperades. Comprova que el fitxer segueixi les dades com el excel a descarregar.")
                    return render(request, "degree_upload_form.html", {"form": form})


                #skip header
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    id_degree = row[0]
                    name_degree = row[1]
                    code_degree = row[2]
                    school_name = row[3]
                    is_active = row[4]

                    if not all([id_degree, name_degree, code_degree]):
                        messages.warning(
                            request, f"Fila {row_num} no s'ha processat correctament: informació incompleta."
                        )
                        error_occurred=True
                        continue

                    if not isinstance(code_degree, (int, float)) or not str(code_degree).isdigit():
                        messages.warning(
                            request, f"Fila {row_num} no s'ha processat: el codi de titulació ha de ser un número."
                        )
                        error_occurred = True
                        continue

                    try:
                        school = School.objects.get(NameSchool=school_name)
                    except School.DoesNotExist:
                        messages.warning(
                            request, f"Fila {row_num} no s'ha processat: l'escola '{school_name}' no existeix."
                        )
                        error_occurred=True
                        continue

                    # Check if the combination of NameDegree and School already exists
                    existing_degree = Degree.objects.filter(idDegree=id_degree).first()

                    if existing_degree:
                        if Degree.objects.filter(NameDegree=name_degree, School=school).exclude(idDegree=existing_degree.idDegree).exists():
                            messages.warning(
                                request, f"Fila {row_num} no s'ha processat: la combinació de titulació i escola ja existeix."
                            )
                            error_occurred = True
                            continue
                        # Update the existing degree
                        existing_degree.CodeDegree = code_degree
                        existing_degree.isActive = is_active.lower() == "actiu" if isinstance(is_active, str) else bool(is_active)
                        existing_degree.save()

                    else:
                        if Degree.objects.filter(NameDegree=name_degree, School=school).exists():
                            # If the combination exists, do not create the degree and warn
                            messages.warning(
                                request, f"Fila {row_num} no s'ha processat: la combinació de titulació i escola ja existeix."
                            )
                            error_occurred = True
                            continue

                        # Create a new degree
                        degree = Degree.objects.create(
                            idDegree=id_degree,
                            NameDegree=name_degree,
                            CodeDegree=code_degree,
                            School=school,
                            isActive=is_active.lower() == "actiu" if isinstance(is_active, str) else bool(is_active),
                        )
                
                if not error_occurred:
                    messages.success(request, "Els graus s'han actualitzat correctament.")
                    return redirect('degree_list')

            except Exception as e:
                messages.error(request, f"Error en processar el fitxer: {e}")
                error_occurred=True

            return render(request, "degree_upload_form.html", {"form": form})

    else:
        form = DegreeUploadForm()

    return render(request, "degree_upload_form.html", {"form": form})